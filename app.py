from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
import random

# Open the Excel file
excel_file = r'zeker.xlsx'  # Replace with your Excel file's path
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active  # Assuming you want to use the active sheet

# Get the number of rows in the Excel sheet
num_rows = sheet.max_row

# Choose a random row (1 to num_rows)
random_row = random.randint(1, num_rows)

# Assuming the text is in column A, you can change this to the actual column
random_text = sheet.cell(row=random_row, column=2).value

# Close the Excel file
workbook.close()

# Now you can use the random_text in your Selenium script


# Create ChromeOptions
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--headless")  # Set headless mode
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")  # This may be required in some environments

# User credentials
username = '01098455410'
password = 'Mustafa_m0512'

# Create a Chrome WebDriver instance with the specified options
driver = webdriver.Chrome(options=chrome_options)

# Navigate to the Facebook login page
driver.get("https://www.facebook.com/messages/t/1840951729334287")

# Locate the email and password input fields by name attribute
email_input = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="email"]')))
password_input = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pass"]')))

# Enter the username and password
email_input.send_keys(username)
password_input.send_keys(password)

# Locate and click the login button
login_button = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.NAME, 'login')))
login_button.click()
time.sleep(15)
# Check if the login was successful by looking for an element unique to logged-in users
if "حتااااهاااااوش" in driver.page_source:
    print('Login successful')
    sender = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[1]/div/div[3]/div/div/div/div[1]/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div/div[2]/div/div/div[4]/div[2]/div/div/div[1]/p")))
    sender.send_keys(random_text)
    sender.send_keys(Keys.ENTER)
    time.sleep(10)

else:
    print('Login failed')


# Close the browser when done
driver.quit()
